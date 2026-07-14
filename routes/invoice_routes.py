from io import StringIO
import pandas as pd
from fastapi import APIRouter, HTTPException, UploadFile, File, Query, Form
from typing import Optional
from models.invoice import FacturaUpdate
from services import invoice_service
from services import report_service
from utils.helpers import clean_nan_values
from fastapi.responses import FileResponse
import os
import tempfile
from openpyxl import load_workbook
from fastapi.responses import StreamingResponse
from io import BytesIO
from services.scheduler_service import last_auto_sync

router = APIRouter(prefix="/invoices", tags=["invoices"])


@router.get("/last-sync")
def get_last_sync():
    """Retorna la fecha/hora de la ultima sincronizacion (manual o automatica)."""
    from config.database import db
    sync_record = db["sync_log"].find_one(sort=[("timestamp", -1)])
    if sync_record:
        return {
            "timestamp": sync_record["timestamp"],
            "type": sync_record.get("type", "unknown"),
            "status": sync_record.get("status", "unknown"),
        }
    return {"timestamp": None, "type": None, "status": None}


@router.get("/dashboard")
def get_invoices_dashboard(
    responsable: Optional[str] = None,
    producto: Optional[str] = None,
    trimestre: Optional[int] = None,
    anio: Optional[int] = None,
):
    try:
        facturas = invoice_service.get_invoices_dashboard(
            responsable=responsable,
            producto=producto,
            trimestre=trimestre,
            anio=anio
        )
        return clean_nan_values(facturas)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Error en get_invoices_dashboard: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al procesar datos: {str(e)}"
        )


@router.get("/filtros")
def get_facturas_filtros():
    try:
        filtros = invoice_service.get_facturas_filtros()
        return filtros
    except Exception as e:
        print(f"Error en get_facturas_filtros: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Error al obtener filtros: {str(e)}")


@router.get("/export_report")
def export_report(segmento: Optional[str] = Query(None, description="Segmento: logistica, facturacion, o vacío para completo")):
    from openpyxl.styles import Font, PatternFill
    from config.database import invoices_collection, vendedores_collection
    import datetime

    try:
        file_path = 'reporte.xlsx'

        if not os.path.exists(file_path):
            raise HTTPException(
                status_code=404, detail="Archivo no encontrado")

        wb = load_workbook(file_path)

        # Renombrar hojas base (no umbral, esas se regeneran)
        rename_map = {
            "Hoja1": "Resumen",
            "Hoja2": "Monto ERP = Excel",
            "Hoja3": "Responsable B24 vs Excel",
            "Hoja4": "OPCI Responsable Único",
            "Hoja5": "Servicios Responsable Fredy",
            "Hoja6": "Nota Crédito Compensada",
        }
        for old_name, new_name in rename_map.items():
            if old_name in wb.sheetnames:
                wb[old_name].title = new_name

        # Eliminar hojas de umbral viejas (Hoja7-10 o Umbral Q*)
        for name in list(wb.sheetnames):
            if name.startswith('Hoja7') or name.startswith('Hoja8') or \
               name.startswith('Hoja9') or name.startswith('Hoja10') or \
               name.startswith('Umbral Q'):
                wb.remove(wb[name])

        # Regenerar hojas de umbral desde MongoDB
        vendedores_db = list(vendedores_collection.find(
            {'esLider': {'$exists': False}}))
        umbrales = {v['nombre']: v['umbral_mensual'] for v in vendedores_db}
        uns = {v['nombre']: v['unidad_negocio'] for v in vendedores_db}
        vendedores_unau = [v['nombre'] for v in vendedores_db if v.get('unidad_negocio') == 'UNAU']

        # Obtener facturas Endress desde MongoDB
        facturas_endress = list(invoices_collection.find({
            "producto_crm": {"$regex": "Endress", "$options": "i"}
        }, {"responsables": 1, "mes": 1, "monto_total": 1}))

        # Construir montos por responsable principal y mes
        endress_data = {}
        for f in facturas_endress:
            responsables = f.get('responsables', [])
            if not responsables:
                continue
            r1 = responsables[0].get('nombre', '')
            mes = f.get('mes')
            monto = f.get('monto_total', 0)
            key = (r1, mes)
            endress_data[key] = endress_data.get(key, 0) + monto

        NOMBRES_MESES = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr',
            5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago',
            9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }

        trimestres = [
            ('Umbral Q1', 'Q1 (Ene–Mar)', [1, 2, 3]),
            ('Umbral Q2', 'Q2 (Abr–Jun)', [4, 5, 6]),
            ('Umbral Q3', 'Q3 (Jul–Sep)', [7, 8, 9]),
            ('Umbral Q4', 'Q4 (Oct–Dic)', [10, 11, 12]),
        ]

        mes_actual = datetime.datetime.now().month

        for sheet_name, q_label, q_meses in trimestres:
            if q_meses[0] > mes_actual:
                continue

            ws = wb.create_sheet(sheet_name)

            # Título
            ws['A1'].value = f'Resumen Umbral Trimestral – {q_label}'
            ws['A1'].font = Font(name='Tahoma', size=11, bold=True)

            # Tabla 1: montos por mes
            T1_ROW = 3
            t1_headers = ['Vendedor'] + [NOMBRES_MESES[m] for m in q_meses] + ['Total General']
            for ci, hdr in enumerate(t1_headers, 1):
                cell = ws.cell(row=T1_ROW, column=ci)
                cell.value = hdr
                cell.fill = PatternFill(start_color='5B9BD5', fill_type='solid')
                cell.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

            for ri, vendedor_nombre in enumerate(vendedores_unau, 1):
                r = T1_ROW + ri
                ws.cell(row=r, column=1).value = vendedor_nombre
                ws.cell(row=r, column=1).font = Font(name='Tahoma', size=9)

                total_general = 0
                for ci, m in enumerate(q_meses, 2):
                    monto = endress_data.get((vendedor_nombre, m), 0)
                    total_general += monto
                    cell = ws.cell(row=r, column=ci)
                    cell.value = monto
                    cell.number_format = '#,##0.00'
                    cell.font = Font(name='Tahoma', size=9)

                tc = ws.cell(row=r, column=len(q_meses) + 2)
                tc.value = total_general
                tc.number_format = '#,##0.00'
                tc.font = Font(name='Tahoma', size=9, bold=True)

            # Tabla 2: umbral y resultado
            T2_ROW = T1_ROW + len(vendedores_unau) + 3

            sub = ws.cell(row=T2_ROW - 1, column=1)
            sub.value = '¿Pasó el umbral?'
            sub.font = Font(name='Tahoma', size=9, bold=True, italic=True)

            t2_headers = ['Vendedor', 'Umbral Mensual', 'Umbral Trimestral', 'Total', '¿Pasó?']
            t2_colors = ['A5A5A5', 'FFC000', 'FFC000', '5B9BD5', '70AD47']
            for ci, (hdr, color) in enumerate(zip(t2_headers, t2_colors), 1):
                cell = ws.cell(row=T2_ROW, column=ci)
                cell.value = hdr
                cell.fill = PatternFill(start_color=color, fill_type='solid')
                cell.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

            for ri, vendedor_nombre in enumerate(vendedores_unau, 1):
                r = T2_ROW + ri
                umbral_m = umbrales.get(vendedor_nombre, 0)
                umbral_t = umbral_m * 3

                total_general = sum(
                    endress_data.get((vendedor_nombre, m), 0)
                    for m in q_meses
                )

                paso = total_general > umbral_t or any(
                    endress_data.get((vendedor_nombre, m), 0) > umbral_m
                    for m in q_meses
                )

                ws.cell(row=r, column=1).value = vendedor_nombre
                ws.cell(row=r, column=1).font = Font(name='Tahoma', size=9)

                for ci, val in [(2, umbral_m), (3, umbral_t), (4, total_general)]:
                    cell = ws.cell(row=r, column=ci)
                    cell.value = val
                    cell.number_format = '#,##0.00'
                    cell.font = Font(name='Tahoma', size=9)

                paso_cell = ws.cell(row=r, column=5)
                paso_cell.value = 'SI' if paso else 'NO'
                paso_cell.fill = PatternFill(
                    start_color='70AD47' if paso else 'FF4040', fill_type='solid')
                paso_cell.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4

        # Filtrar hojas según segmento
        SEGMENTOS = {
            "logistica": {
                "Responsable B24 vs Excel",
                "OPCI Responsable Único",
                "Servicios Responsable Fredy",
                "Datos Incompletos",
                "Resumen Validaciones",
            },
            "facturacion": {
                "Monto ERP = Excel",
                "Nota Crédito Compensada",
                "Resumen Validaciones",
            },
        }

        if segmento and segmento in SEGMENTOS:
            keep = SEGMENTOS[segmento]
            for name in list(wb.sheetnames):
                if name not in keep:
                    wb.remove(wb[name])

        suffix = f"_{segmento}" if segmento and segmento in SEGMENTOS else ""

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=reporte_invoices{suffix}.xlsx"}
        )
    except Exception as e:
        print(f"Error en export_report: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error al exportar archivo: {str(e)}")


@router.post("/execute_report")
async def get_facturas_filtros(
    file: Optional[UploadFile] = File(None),
    ventas_data: Optional[str] = Form(None)
):
    try:
        df_invoices = None
        if file:
            contents = await file.read()
            df_invoices = pd.read_csv(
                StringIO(contents.decode('utf-8')),
                sep=';'
            )

        df_ventas = None
        if ventas_data:
            import json
            ventas_json = json.loads(ventas_data)
            df_ventas = pd.DataFrame(ventas_json)

        report_service.execute_report(
            data_invoices=df_invoices,
            data_ventas=df_ventas
        )
        if file and ventas_data:
            msg = 'Actualización exitosa con archivo CSV y datos de BD'
        elif file:
            msg = 'Actualización exitosa con archivo CSV'
        else:
            msg = 'Actualización exitosa con datos de BD local'
        return {'message': msg}

    except Exception as e:
        print(f"Error en get_facturas_filtros: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al procesar: {str(e)}"
        )


@router.get("/execute_report_by_user")
async def get_facturas_by_user(name_user: str = Query(..., description="Nombre del usuario")):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl import Workbook
    from config.database import invoices_collection, vendedores_collection

    try:
        # Leer directamente de MongoDB para tener datos actualizados
        facturas = list(invoices_collection.find(
            {"responsables.nombre": name_user}
        ))

        if not facturas:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontraron datos para: {name_user}"
            )

        # Obtener info del vendedor desde la BD
        vendedor = vendedores_collection.find_one({'nombre': name_user})

        # Extraer responsables del array para columnas planas
        rows = []
        for f in facturas:
            responsables = f.get('responsables', [])
            r1 = responsables[0] if len(responsables) >= 1 else {}
            r2 = responsables[1] if len(responsables) >= 2 else {}

            comision_total = f.get('comision_total', 0)

            rows.append({
                'OK': f.get('ok', ''),
                'AÑO': f.get('anio'),
                'MES': f.get('mes'),
                'Responsable 1': r1.get('nombre', ''),
                'Responsable 2': r2.get('nombre', ''),
                'Unidad de Negocio': f.get('unidad_negocio', ''),
                'Fecha': f.get('fecha'),
                'Estado': f.get('estado', ''),
                'Número': f.get('numero', ''),
                'Monto Total': f.get('monto_total', 0),
                'Producto CRM': f.get('producto_crm', ''),
                'UBruta': f.get('utilidad_bruta', 0),
                'Nombre Empresa': f.get('nombre_empresa', ''),
                'Subject': f.get('subject', ''),
                'Codigos': f.get('codigos', ''),
                'Cotizacion #': f.get('cotizacion_num', ''),
                'Proviene EPC/OEM/Canal Deal?': f.get('origen_deal', ''),
                'T/C de la Factura': f.get('tipo_cambio_factura', 0),
                'Monto Actualizado': f.get('monto_actualizado', 0),
                'Diferencia': f.get('diferencia', ''),
                'Notas': f.get('notas', ''),
                'Observaciones': f.get('observaciones', ''),
                'Periodo': f.get('periodo', ''),
                'EstadoPago-Vendedor': f.get('estado_pago_vendedor', ''),
                'Lider 1': f.get('lider_1', ''),
                'Lider 2': f.get('lider_2', ''),
                'EstadoPago-Lideres': f.get('estado_pago_lideres', ''),
                'Umbral': f.get('umbral', 0),
                'Comisiona': f.get('comisiona', True),
                'Comisión 1': r1.get('comision', 0),
                'Comisión 2': r2.get('comision', 0),
            })

        df = pd.DataFrame(rows)

        # Columnas del reporte
        columnas = [
            'OK', 'AÑO', 'MES', 'Responsable 1', 'Responsable 2',
            'Unidad de Negocio', 'Fecha', 'Estado', 'Número', 'Monto Total',
            'Producto CRM', 'UBruta', 'Nombre Empresa', 'Subject', 'Codigos',
            'Cotizacion #', 'Proviene EPC/OEM/Canal Deal?', 'T/C de la Factura',
            'Monto Actualizado', 'Diferencia', 'Notas', 'Observaciones',
            'Periodo', 'EstadoPago-Vendedor', 'Lider 1', 'Lider 2',
            'EstadoPago-Lideres', 'Umbral', 'Comisiona', 'Comisión 1', 'Comisión 2'
        ]
        df = df[columnas]

        # Crear Excel con formato
        wb = Workbook()
        ws = wb.active
        ws.title = 'Resumen'

        # Headers
        for ci, col_name in enumerate(columnas, 1):
            ws.cell(row=1, column=ci).value = col_name

        # Datos
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, col_name in enumerate(columnas, 1):
                val = row[col_name]
                if pd.isna(val) if isinstance(val, float) else False:
                    val = ''
                ws.cell(row=ri, column=ci).value = val

        num_filas = len(df)

        # Formato de celdas de datos
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Tahoma", size=9)

        # Colores de cabecera (mismos que reporte completo)
        header_colors = [
            [1, 3, 'A5A5A5'],
            [4, 18, '5B9BD5'],
            [19, 21, 'A5A5A5'],
            [22, 22, '92D050'],
            [23, 28, 'A5A5A5'],
            [29, 31, '70AD47'],
        ]
        for inicio, fin, color in header_colors:
            for col in range(inicio, min(fin + 1, len(columnas) + 1)):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color=color, fill_type="solid")
                cell.font = Font(name="Tahoma", size=9, bold=True, color="FFFFFFFF")

        # Formatos numéricos
        for i in range(2, num_filas + 2):
            cell = ws.cell(row=i, column=12)  # UBruta (L)
            if (cell.value or 0) >= 0.22:
                cell.number_format = '">"0%'
            else:
                cell.number_format = "0.00%"
            cell.fill = PatternFill(start_color="A9D08E", fill_type="solid")
            cell.font = Font(color="375623")

            ws.cell(row=i, column=19).number_format = "0.00"  # Monto Actualizado (S)
            ws.cell(row=i, column=18).number_format = "0.00"  # T/C (R)
            ws.cell(row=i, column=7).number_format = 'DD/MM/YYYY'  # Fecha (G)
            ws.cell(row=i, column=30).number_format = "0.00"  # Comisión 1
            ws.cell(row=i, column=31).number_format = "0.00"  # Comisión 2

        # Formato condicional: montos negativos en rojo
        ws.conditional_formatting.add(
            f"S2:S{num_filas + 1}",
            FormulaRule(formula=["S2<0"], font=Font(color="FF0000"))
        )
        ws.conditional_formatting.add(
            f"J2:J{num_filas + 1}",
            FormulaRule(
                formula=["J2<0"],
                font=Font(color="FF0000"),
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            )
        )

        # Auto-ancho de columnas
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        ws.auto_filter.ref = ws.dimensions

        # Hoja de umbral Endress (si es UNAU)
        if vendedor and vendedor.get('unidad_negocio') == 'UNAU':
            NOMBRES_MESES = {
                1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr',
                5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago',
                9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
            }
            trimestres = [
                ('Q1 (Ene-Mar)', [1, 2, 3]),
                ('Q2 (Abr-Jun)', [4, 5, 6]),
                ('Q3 (Jul-Sep)', [7, 8, 9]),
                ('Q4 (Oct-Dic)', [10, 11, 12]),
            ]

            # Facturas Endress del usuario
            facturas_endress = list(invoices_collection.find({
                "responsables.nombre": name_user,
                "producto_crm": {"$regex": "Endress", "$options": "i"}
            }))

            umbral_mensual = vendedor.get('umbral_mensual', 0)
            umbral_trimestral = vendedor.get('umbral_trimestral', 0)

            import datetime
            mes_actual = datetime.datetime.now().month

            for q_label, q_meses in trimestres:
                if q_meses[0] > mes_actual:
                    continue

                ws_q = wb.create_sheet(f'Umbral {q_label[:2]}')

                # Título
                ws_q['A1'].value = f'Umbral Endress – {q_label} – {name_user}'
                ws_q['A1'].font = Font(name='Tahoma', size=11, bold=True)

                # Tabla de montos por mes
                headers = ['Mes', 'Monto', 'Umbral Mensual', '¿Pasó?']
                for ci, h in enumerate(headers, 1):
                    cell = ws_q.cell(row=3, column=ci)
                    cell.value = h
                    cell.fill = PatternFill(start_color='5B9BD5', fill_type='solid')
                    cell.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

                total_trimestre = 0
                for ri, mes in enumerate(q_meses, 4):
                    monto_mes = sum(
                        f.get('monto_total', 0) for f in facturas_endress
                        if f.get('mes') == mes
                    )
                    total_trimestre += monto_mes
                    paso_mes = monto_mes > umbral_mensual

                    ws_q.cell(row=ri, column=1).value = NOMBRES_MESES[mes]
                    ws_q.cell(row=ri, column=1).font = Font(name='Tahoma', size=9)

                    cell_m = ws_q.cell(row=ri, column=2)
                    cell_m.value = monto_mes
                    cell_m.number_format = '#,##0.00'
                    cell_m.font = Font(name='Tahoma', size=9)

                    cell_u = ws_q.cell(row=ri, column=3)
                    cell_u.value = umbral_mensual
                    cell_u.number_format = '#,##0.00'
                    cell_u.font = Font(name='Tahoma', size=9)

                    cell_p = ws_q.cell(row=ri, column=4)
                    cell_p.value = 'SI' if paso_mes else 'NO'
                    cell_p.fill = PatternFill(
                        start_color='70AD47' if paso_mes else 'FF4040', fill_type='solid')
                    cell_p.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

                # Fila total
                r_total = 4 + len(q_meses)
                paso_trim = total_trimestre > umbral_trimestral

                ws_q.cell(row=r_total, column=1).value = 'TOTAL'
                ws_q.cell(row=r_total, column=1).font = Font(name='Tahoma', size=9, bold=True)

                cell_t = ws_q.cell(row=r_total, column=2)
                cell_t.value = total_trimestre
                cell_t.number_format = '#,##0.00'
                cell_t.font = Font(name='Tahoma', size=9, bold=True)

                cell_ut = ws_q.cell(row=r_total, column=3)
                cell_ut.value = umbral_trimestral
                cell_ut.number_format = '#,##0.00'
                cell_ut.font = Font(name='Tahoma', size=9, bold=True)

                cell_pt = ws_q.cell(row=r_total, column=4)
                cell_pt.value = 'SI' if paso_trim else 'NO'
                cell_pt.fill = PatternFill(
                    start_color='70AD47' if paso_trim else 'FF4040', fill_type='solid')
                cell_pt.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

                for col in ws_q.columns:
                    max_len = max((len(str(c.value or '')) for c in col), default=10)
                    ws_q.column_dimensions[col[0].column_letter].width = max_len + 4

        # Guardar y enviar
        temp_dir = tempfile.gettempdir()
        temp_filename = f"reporte_{name_user.replace(' ', '_')}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_path = os.path.join(temp_dir, temp_filename)
        wb.save(temp_path)
        wb.close()

        return FileResponse(
            path=temp_path,
            filename=f"reporte_{name_user.replace(' ', '_')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/analisis")
def get_analisis():
    try:
        from services.analisis_service import get_analisis as _get_analisis
        return _get_analisis()
    except Exception as e:
        print(f"Error en get_analisis: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al obtener análisis: {str(e)}")


@router.post("/sync-bitrix")
async def sync_from_bitrix(ventas_data: Optional[str] = Form(None)):
    from config.database import db
    from datetime import datetime, timezone

    try:
        from services.bitrix_service import fetch_invoices_from_bitrix
        import json as _json

        df_invoices = fetch_invoices_from_bitrix()

        df_ventas = None
        if ventas_data:
            df_ventas = pd.DataFrame(_json.loads(ventas_data))

        report_service.execute_report(
            data_invoices=df_invoices,
            data_ventas=df_ventas,
        )

        db["sync_log"].insert_one({
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "type": "manual",
            "status": "success",
            "message": f"{len(df_invoices)} invoices sincronizadas",
        })

        return {'message': f'Sincronización exitosa: {len(df_invoices)} invoices obtenidas de Bitrix24'}

    except Exception as e:
        print(f"Error en sync_from_bitrix: {str(e)}")
        import traceback
        traceback.print_exc()
        db["sync_log"].insert_one({
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "type": "manual",
            "status": "error",
            "message": str(e),
        })
        raise HTTPException(status_code=500, detail=f"Error al sincronizar con Bitrix24: {str(e)}")


@router.get("")
def get_all_facturas(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    producto: Optional[str] = None,
    responsable: Optional[str] = None,
    mes: Optional[int] = None,
    anio: Optional[int] = None,
):
    try:
        result = invoice_service.get_all_facturas(
            skip, limit, search, producto, responsable, mes, anio
        )
        result['facturas'] = clean_nan_values(result['facturas'])
        return result
    except Exception as e:
        print(e)
        raise HTTPException(
            status_code=500, detail=f"Error al obtener facturas: {str(e)}")
