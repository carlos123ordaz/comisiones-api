import pandas as pd
from msal import ConfidentialClientApplication
import requests
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
import numpy as np
import unicodedata
from typing import Optional
from config import database
from dotenv import load_dotenv
load_dotenv()


def execute_report(
    data_invoices: Optional[pd.DataFrame] = None,
    data_ventas: Optional[pd.DataFrame] = None
):
    TENANT_ID = os.getenv("TENANT_ID")
    CLIENT_ID = os.getenv("CLIENT_ID")
    CLIENT_SECRET = os.getenv("CLIENT_SECRET")

    app = ConfidentialClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        client_credential=CLIENT_SECRET
    )
    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"])

    if "access_token" not in result:
        raise ValueError(
            f"Error de autenticación con Microsoft 365: {result.get('error_description', result.get('error', 'Token no obtenido'))}. "
            "Renueva el CLIENT_SECRET en el portal de Azure AD."
        )

    token = result["access_token"]
    headers = {'Authorization': f'Bearer {token}'}

    site_url = "https://graph.microsoft.com/v1.0/sites/corsusaadmin.sharepoint.com:/sites/logistica"
    site_response = requests.get(site_url, headers=headers)
    site_id = site_response.json()['id']

    drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    drives_response = requests.get(drives_url, headers=headers)
    drives = drives_response.json()['value']

    document_drive = None
    for d in drives:
        if 'Documentos' in d['name'] or 'Documents' in d['name']:
            document_drive = d
            break

    if not document_drive:
        document_drive = drives[0]

    drive_id = document_drive['id']
    archivos_config = [
        {
            'unique_id': 'F4A1F287-E4E8-4F04-B39D-215DC433A483',
            'nombre': '001_Ventas_OP.xlsx'
        },
        {
            'unique_id': '8C6018D0-2120-4457-89EB-ED641E69B3EA',
            'nombre': '004_Facturacion_OP.xlsx'
        }
    ]

    download_dir = './descargas'
    os.makedirs(download_dir, exist_ok=True)

    def descargar_archivo(config, index, total):
        unique_id = config['unique_id']
        nombre = config['nombre']
        file_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{unique_id}"
        file_response = requests.get(file_url, headers=headers)

        if file_response.status_code != 200:
            return None, f"Error al obtener info del archivo {nombre}"

        file_data = file_response.json()
        download_url = file_data.get('@microsoft.graph.downloadUrl')
        file_content = requests.get(download_url)
        ruta_archivo = os.path.join(download_dir, nombre)

        with open(ruta_archivo, 'wb') as f:
            f.write(file_content.content)

        return ruta_archivo, f"Archivo {nombre} descargado exitosamente"

    for idx, config in enumerate(archivos_config, 1):
        ruta, mensaje = descargar_archivo(config, idx, len(archivos_config))
        print(mensaje)

    a = pd.read_excel('./descargas/001_Ventas_OP.xlsx',
                      sheet_name='OP_Cabecera', skiprows=2)
    b = pd.read_excel('./descargas/004_Facturacion_OP.xlsx', skiprows=3)
    c = pd.read_excel('./descargas/001_Ventas_OP.xlsx',
                      sheet_name='OP_Detalle-Venta', skiprows=2)

    if data_invoices is not None:
        bt = data_invoices
        bt.to_csv('./descargas/invoices.csv', sep=';', index=False)
    else:
        bt = pd.read_csv('./descargas/invoices.csv', sep=';')

    num_deal = {}
    for _, row in c.iterrows():
        num_deal[row['Correlativo_OPCI']] = row['Numero_Deal']

    f = data_ventas

    bt['Factura #'] = bt['Factura #'].str.split(' ').str[0]
    un = {}
    for index, row in bt.iterrows():
        un[row['Factura #']] = row['Unidad de Negocio']

    f['Num_Factura'] = f['NroSre'] + '-' + f['NroDoc']
    f['ValorNeto'] = np.where(
        f['Cd_Mda'] == '01', f['ValorNeto']/f['CamMda'], f['ValorNeto'])
    b['Monto'] = np.where(b['Moneda'] == 'PEN', b['MontoTotal_SinIGV'] /
                          b['T/C_USD-Sol'], b['MontoTotal_SinIGV'])
    grupo = b.groupby('Num_Factura').agg({
        'Monto': 'sum',
        'Status_Factura': 'first',
        'T/C_USD-Sol': 'first'
    }).reset_index()
    one = pd.merge(f, grupo, on='Num_Factura', how='inner')
    one = one[['Num_Factura', 'ValorNeto', 'Monto',
               'Status_Factura', 'CamMda', 'T/C_USD-Sol']]
    one['Monto (-)'] = round(abs(abs(one['ValorNeto']) - abs(one['Monto'])), 2)
    one['T/C (-)'] = round(abs(one['CamMda'] - one['T/C_USD-Sol']), 2)
    one = one.rename(columns={'Num_Factura': 'Número', 'ValorNeto': 'Monto ERP',
                     'Monto': 'Monto Excel', 'CamMda': 'T/C ERP', 'T/C_USD-Sol': 'T/C Excel'})
    bt['OPCI'] = 'OPCI-' + \
        bt['Nombre'].str.extract(r'OPCI-(\d+)', expand=False)
    obj_r1 = {}
    obj_r2 = {}
    for index, row in bt.iterrows():
        obj_r1[row['OPCI']] = row['Responsable Deal - Principal']
        obj_r2[row['OPCI']] = row['Responsable Deal - Secundario']

    nc1 = f[f['Cd_TD'] == '07'].copy()
    b = b[~b['Num_Factura'].isin(nc1['Num_Factura'])]
    df = pd.merge(f, b, how='left', on='Num_Factura')
    df = pd.merge(df, a, how='left', on='Correlativo_OPCI')
    df['UN'] = df['Num_Factura'].map(un)
    df['FecMov'] = pd.to_datetime(df['FecMov'], errors='coerce')
    df['AÑO'] = df['FecMov'].dt.year
    df['MES'] = df['FecMov'].dt.month
    df['Subject'] = '-'
    df['Codigos'] = '-'
    df['Diferencia'] = '-'
    df['Notas'] = '-'
    df['Observaciones'] = '-'
    df['Periodo'] = df['AÑO'].astype(
        str) + '-Q' + ((df['MES'] - 1) // 3 + 1).astype(str)
    df['EstadoPago-Vendedor'] = ''
    df['Lider 2'] = ''
    df['EstadoPago-Lideres'] = ''
    df['Umbral'] = 0.22
    df['Invoice items: Name'] = '-'
    df['OK'] = ''
    df['Cotizacion #'] = df['Correlativo_OPCI']
    df['deal'] = df['Correlativo_OPCI'].map(num_deal)
    df['Subject'] = df['CA10']
    df['Monto Total'] = np.where(
        df['Moneda_x'] == 'USD',
        df['MontoTotal_SinIGV_x'],
        df['MontoTotal_SinIGV_x'] / df['CamMda']
    )
    df['Monto Total'] = np.where(
        df['Cd_TD'] == '07',
        -abs(df['Monto Total']),
        df['Monto Total']
    )
    df['Estado'] = np.where(df['Cd_TD'] == '07',
                            'Nota Crédito', df['Status_Factura'])
    df['R1'] = df['Correlativo_OPCI'].map(obj_r1)
    df['R2'] = df['Correlativo_OPCI'].map(obj_r2)
    nc = df[df['Cd_TD'] == '07'].copy()
    nc['REF'] = nc['DR_NSre'] + '-' + nc['DR_NDoc']
    ref_nc = {}
    umb = {}
    for index, row in df.iterrows():
        umb[row['Num_Factura']] = row['UBrutaCoti']

    ref_umb = {}
    for index, row in nc.iterrows():
        ref_nc[row['REF']] = row['Num_Factura']
        ref_umb[row['Num_Factura']] = umb.get(row['REF'])
    df['REF'] = df['Num_Factura'].map(ref_nc)
    df['REF_2'] = df['Num_Factura'].map(ref_umb)
    df['UBrutaCoti'] = np.where(
        df['Cd_TD'] == '07', df['REF_2'], df['UBrutaCoti'])
    df['Estado'] = np.where(
        df['REF'].notna() & (df['REF'] != ''),
        'Factura - ' + df['REF'].astype(str),
        df['Estado']
    )

    df['Monto Total'] = np.where(
        df['Estado'].str.contains('Nota Crédito', na=False),
        -abs(df['ValorNeto']),
        df['Monto Total']
    )

    df['Monto Total'] = np.where(
        df['Estado'].isna() |
        (df['Estado'] == '') |
        (df['Estado'].str.contains('Factura', na=False) & df['Monto Total'].isna()),
        df['ValorNeto'],
        df['Monto Total']
    )

    r1 = {}
    r2 = {}
    un = {}
    l1 = {}
    for index, row in df.iterrows():
        v1 = row['Vendedor1']
        if (row['CA10'] not in r1) and pd.notna(v1) and v1 != "":
            r1[row['CA10']] = v1
        v2 = row['Vendedor2']
        if (row['CA10'] not in r2) and pd.notna(v2) and v2 != "":
            r2[row['CA10']] = v2
        n1 = row['UN']
        if (row['CA10'] not in un) and pd.notna(n1) and n1 != "":
            un[row['CA10']] = n1
        l = row['Lider']
        if (row['CA10'] not in l1) and pd.notna(l) and l != "":
            l1[row['CA10']] = l

    df['Responsable 1_A'] = df['CA10'].map(r1)
    df['Responsable 2_A'] = df['CA10'].map(r2)
    df['Unidad de Negocio_A'] = df['CA10'].map(un)
    df['Lider_A'] = df['CA10'].map(l1)
    df['Lider'] = np.where(df['Lider'].isna(), df['Lider_A'], df['Lider'])
    df['Responsable 1'] = np.where(
        df['Vendedor1'].isna(), df['Responsable 1_A'], df['Vendedor1'])
    df['Responsable 2'] = np.where(
        df['Vendedor2'].isna(), df['Responsable 2_A'], df['Vendedor2'])
    df['Unidad de Negocio'] = np.where(
        df['UN'].isna(), df['Unidad de Negocio_A'], df['UN'])
    df['Proviene EPC/OEM/Canal Deal?'] = np.where(
        df['Responsable 1'] == df['Responsable 2'], 'No', 'Si')
    df['UBrutaCoti'] = df['UBrutaCoti'].fillna(0)
    es_servicio = df['Producto_CRM'].str.contains('Serv-', na=False)
    df['UBrutaCoti'] = np.where(
        es_servicio & (df['UBrutaCoti'] < 0.22), 0.22, df['UBrutaCoti'])
    df['UBrutaCoti'] = np.where(
        df['UBrutaCoti'] >= 0.22, 0.22, df['UBrutaCoti'])
    df['Monto Actualizado'] = np.where(
        df['UBrutaCoti'] >= 0.22, df['Monto Total'], df['Monto Total'] * df['UBrutaCoti'] / 0.22)
    df['Producto_CRM'] = df['Producto_CRM'].fillna('-')
    new_column = {
        'FecMov': 'Fecha',
        'Num_Factura': 'Número',
        'Producto_CRM': 'Producto CRM',
        'UBrutaCoti': 'UBruta',
        'Cliente': 'Nombre Empresa',
        'CamMda': 'T/C de la Factura',
        'Lider': 'Lider 1'
    }
    df.rename(columns=new_column, inplace=True)

    df_c1 = df[['AÑO', 'MES', 'Fecha', 'Número', 'Correlativo_OPCI',
                'Responsable 1', 'Responsable 2', 'R1', 'R2']].copy()
    df_c1['Responsable 1'] = df_c1['Responsable 1'].apply(lambda x: unicodedata.normalize(
        'NFD', x).encode('ascii', 'ignore').decode('utf-8') if isinstance(x, str) else x)
    df_c1['Responsable 2'] = df_c1['Responsable 2'].apply(lambda x: unicodedata.normalize(
        'NFD', x).encode('ascii', 'ignore').decode('utf-8') if isinstance(x, str) else x)
    df_c1['R1'] = df_c1['R1'].apply(lambda x: unicodedata.normalize('NFD', x).encode(
        'ascii', 'ignore').decode('utf-8') if isinstance(x, str) else x)
    df_c1['R2'] = df_c1['R2'].apply(lambda x: unicodedata.normalize('NFD', x).encode(
        'ascii', 'ignore').decode('utf-8') if isinstance(x, str) else x)
    df_c1 = df_c1[df_c1['Correlativo_OPCI'].notna()]
    df_c1.rename(columns={'Responsable 1': 'Responsable 1 E.', 'Responsable 2': 'Responsable 2 E.',
                 'R1': 'Responsable 1 B.', 'R2': 'Responsable 2 B.'}, inplace=True)

    grupo_opci = bt.groupby('OPCI')
    conflictos = []

    for opci, grupo in grupo_opci:
        responsables_1 = grupo['Responsable Deal - Principal'].dropna().unique()
        responsables_2 = grupo['Responsable Deal - Secundario'].dropna().unique()

        if len(responsables_1) > 1 or len(responsables_2) > 1:
            conflictos.append(grupo[['OPCI', 'Nombre', 'Factura #', 'Fecha de la factura',
                              'Etapa', 'Responsable Deal - Principal', 'Responsable Deal - Secundario']])

    if conflictos:
        df_conflictos = pd.concat(conflictos, ignore_index=True)
    else:
        df_conflictos = pd.DataFrame(columns=[
            'OPCI', 'Nombre', 'Factura #', 'Fecha de la factura',
            'Etapa', 'Responsable Deal - Principal', 'Responsable Deal - Secundario'
        ])

    df_servicios = df[df['Producto CRM'].str.contains('Serv-', na=False)]
    df_servicios = df_servicios[['AÑO', 'MES', 'Fecha', 'Número', 'Producto CRM',
                                 'Cotizacion #', 'Responsable 1', 'Responsable 2', 'Lider 1']]

    df['Factura Relacionada'] = df['DR_NSre'] + '-' + df['DR_NDoc']
    facturas = set(df['Número'])
    nc = df[df['Cd_TD'] == '07'].copy()
    df_group = (
        df.groupby('Número', as_index=False)
        .agg(Total_Factura=('Monto Total', 'sum'))
    )
    nc = nc.merge(
        df_group,
        left_on='Factura Relacionada',
        right_on='Número',
        how='left',
        suffixes=('', '_Factura')
    )
    nc['Factura encontrada'] = nc['Factura Relacionada'].isin(facturas)
    nc['Diferencia'] = abs(nc['Total_Factura'] + nc['Monto Total'])
    hoja6 = nc[['AÑO', 'MES', 'Estado', 'Número', 'Monto Total', 'Producto CRM', 'Cotizacion #', 'Responsable 1',
                'Responsable 2', 'Lider 1', 'Factura Relacionada', 'Total_Factura', 'Diferencia', 'Factura encontrada']]

    df = df[['OK', 'AÑO', 'MES', 'Responsable 1', 'Responsable 2', 'Unidad de Negocio', 'Fecha', 'Estado', 'Número', 'Monto Total', 'Producto CRM', 'UBruta',
            'Nombre Empresa', 'Subject', 'Codigos', 'Cotizacion #', 'Proviene EPC/OEM/Canal Deal?', 'T/C de la Factura', 'Monto Actualizado', 'Diferencia', 'Notas',
             'Observaciones', 'Periodo', 'EstadoPago-Vendedor', 'Lider 1', 'Lider 2', 'EstadoPago-Lideres', 'Umbral']]

    # Hoja Datos Incompletos: filas donde campos clave son nulos o guion
    _cols_validar = ['Responsable 1', 'Responsable 2', 'UBruta', 'Cotizacion #', 'Producto CRM']
    def _campo_incompleto(val):
        if val is None:
            return True
        s = str(val).strip()
        return s == '' or s == '-' or s.lower() == 'nan'

    _mask = df[_cols_validar].apply(lambda col: col.map(_campo_incompleto)).any(axis=1)
    df_incompletos = df[_mask].drop(columns=['Unidad de Negocio']).copy()
    df_incompletos['Campos Incompletos'] = df[_cols_validar].apply(
        lambda col: col.map(_campo_incompleto)
    ).apply(lambda row: ', '.join(_cols_validar[i] for i, v in enumerate(row) if v), axis=1)

    # Resumen de validaciones
    _n_incompletos = len(df_incompletos)
    _n_monto_erp = len(one[(one['Monto (-)'] > 0) | (one['T/C (-)'] > 0)])
    _n_responsable = len(df_c1[
        (df_c1['Responsable 1 B.'] != df_c1['Responsable 1 E.']) |
        (df_c1['Responsable 2 B.'] != df_c1['Responsable 2 E.'])
    ])
    _n_opci = len(df_conflictos)
    _n_fredy = len(df_servicios[
        (df_servicios['Responsable 1'] != 'Fredy Huaman R.') |
        (df_servicios['Responsable 2'] != 'Fredy Huaman R.')
    ])
    _n_nc = len(hoja6[(hoja6['Diferencia'] > 0) | (hoja6['Factura encontrada'] == False)])

    df_resumen_val = pd.DataFrame([
        {'Validación': 'Datos Incompletos',        'Descripción': 'Facturas con Responsable, Margen, OPCI o Producto vacío o guion', 'Errores': _n_incompletos, 'Archivo': 'Logística'},
        {'Validación': 'Monto ERP ≠ Excel',         'Descripción': 'Diferencia en monto o tipo de cambio entre ERP y Facturacion_OP',    'Errores': _n_monto_erp,   'Archivo': 'Facturación'},
        {'Validación': 'Responsable B24 vs Excel',  'Descripción': 'Responsables distintos entre Bitrix24 y Facturacion_OP',              'Errores': _n_responsable, 'Archivo': 'Logística'},
        {'Validación': 'OPCI Responsable Único',    'Descripción': 'OPCI con múltiples responsables asignados',                       'Errores': _n_opci,        'Archivo': 'Logística'},
        {'Validación': 'Servicios - Resp. Fredy',   'Descripción': 'Servicios sin Fredy Huaman R. como responsable',                  'Errores': _n_fredy,       'Archivo': 'Logística'},
        {'Validación': 'Nota Crédito Compensada',   'Descripción': 'Notas crédito con diferencia pendiente o factura no encontrada',  'Errores': _n_nc,          'Archivo': 'Facturación'},
    ])

    with pd.ExcelWriter('reporte.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Hoja1', index=False)
        one.to_excel(writer, sheet_name='Hoja2', index=False)
        df_c1.to_excel(writer, sheet_name='Hoja3', index=False)
        df_conflictos.to_excel(writer, sheet_name='Hoja4', index=False)
        df_servicios.to_excel(writer, sheet_name='Hoja5', index=False)
        hoja6.to_excel(writer, sheet_name='Hoja6', index=False)
        df_incompletos.to_excel(writer, sheet_name='Datos Incompletos', index=False)
        df_resumen_val.to_excel(writer, sheet_name='Resumen Validaciones', index=False)

    nombre_archivo = "reporte.xlsx"
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    num_filas = len(df)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Tahoma", size=9)

    for i in range(2, num_filas + 2):
        ws[f'T{i}'] = f'=IF(L{i}="",0,S{i}-J{i})'

    items = [
        [1, 3, 'A5A5A5'],
        [4, 18, '5B9BD5'],
        [19, 21, 'A5A5A5'],
        [22, 22, '92D050'],
        [23, 28, 'A5A5A5'],
    ]

    for item in items:
        inicio, fin, color = item
        for col in range(inicio, fin + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=color,  fill_type="solid")
            cell.font = Font(
                name="Tahoma",
                size=9,
                bold=True,
                color="FFFFFFFF"
            )

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 2

    for i in range(2, num_filas + 2):
        cell = ws[f"L{i}"]
        if (cell.value or 0) >= 0.22:
            cell.number_format = '">"0%'
        else:
            cell.number_format = "0.00%"
        cell.fill = PatternFill(start_color="A9D08E", fill_type="solid")
        cell.font = Font(color="375623")
        cell = ws[f"AB{i}"]
        cell.number_format = "0.00%"
        cell = ws[f"S{i}"]
        cell.number_format = "0.00"
        cell = ws[f"T{i}"]
        cell.number_format = "0.00"
        cell = ws[f"R{i}"]
        cell.number_format = "0.00"
        cell = ws[f'G{i}']
        cell.number_format = 'DD/MM/YYYY'

    ws.conditional_formatting.add(
        f"S2:S{num_filas + 1}",
        FormulaRule(formula=[f"S2<0"], font=Font(color="FF0000"))
    )
    ws.conditional_formatting.add(
        f"T2:T{num_filas + 1}",
        FormulaRule(formula=[f"T2<0"], font=Font(color="FF0000"))
    )

    ws.conditional_formatting.add(
        f"L2:L{num_filas + 1}",
        FormulaRule(
            formula=[f"L2=0"],
            font=Font(color="FFFFFF"),
            fill=PatternFill(start_color="FF0000",
                             end_color="FF0000", fill_type="solid")
        )
    )

    ws.conditional_formatting.add(
        f"L2:L{num_filas + 1}",
        FormulaRule(
            formula=[f"L2<0.22"],
            font=Font(color="9C002A"),
            fill=PatternFill(start_color="F8CBAD",
                             end_color="F8CBAD", fill_type="solid")
        )
    )

    ws.conditional_formatting.add(
        f"K2:K{num_filas + 1}",
        FormulaRule(
            formula=[f'COUNTIF(K2,"*Proy-*")>0'],
            fill=PatternFill(start_color="DBDBDB",
                             end_color="DBDBDB", fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        f"K2:K{num_filas + 1}",
        FormulaRule(
            formula=[f'COUNTIF(K2,"*Serv-*")>0'],
            fill=PatternFill(start_color="FFD966",
                             end_color="FFD966", fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        f"J2:J{num_filas + 1}",
        FormulaRule(
            formula=[f'J2<0'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.column_dimensions['K'].width = 28
    ws.column_dimensions['M'].width = 35
    ws.column_dimensions['T'].width = 14
    ws.column_dimensions['S'].width = 23
    ws.column_dimensions['N'].width = 44
    ws.auto_filter.ref = ws.dimensions

    # Hoja N° 2 — Monto ERP = Excel
    # Pre-filtro: ocultar filas sin discrepancia (G<=0 Y H<=0)
    ws = wb["Hoja2"]
    ws.conditional_formatting.add(
        f"G2:G{num_filas + 1}",
        FormulaRule(
            formula=[f'G2>0'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        f"H2:H{num_filas + 1}",
        FormulaRule(
            formula=[f'H2>0'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 2
    for r in range(2, ws.max_row + 1):
        g_val = ws.cell(row=r, column=7).value or 0
        h_val = ws.cell(row=r, column=8).value or 0
        if not (g_val > 0 or h_val > 0):
            ws.row_dimensions[r].hidden = True

    # Hoja N° 3 — Responsable B24 vs Excel
    # Pre-filtro: ocultar filas donde H==F y I==G (sin discrepancia)
    ws = wb["Hoja3"]
    ws.conditional_formatting.add(
        f"H2:H{num_filas + 1}",
        FormulaRule(
            formula=[f'H2<>F2'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        f"I2:I{num_filas + 1}",
        FormulaRule(
            formula=[f'I2<>G2'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 2
    for r in range(2, ws.max_row + 1):
        f_val = ws.cell(row=r, column=6).value
        g_val = ws.cell(row=r, column=7).value
        h_val = ws.cell(row=r, column=8).value
        i_val = ws.cell(row=r, column=9).value
        if h_val == f_val and i_val == g_val:
            ws.row_dimensions[r].hidden = True

    # Hoja N° 4
    ws = wb["Hoja4"]
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 2

    color1 = PatternFill(start_color="DAEEF3",
                         end_color="DAEEF3", fill_type="solid")
    color2 = PatternFill(start_color="FDE9D9",
                         end_color="FDE9D9", fill_type="solid")
    grupo_actual = None
    color_actual = color1
    alternar = True

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        valor_grupo = row[0].value
        if valor_grupo != grupo_actual:
            grupo_actual = valor_grupo
            color_actual = color1 if alternar else color2
            alternar = not alternar
        for cell in row:
            cell.fill = color_actual

    # Hoja N° 5 — Servicios Responsable Fredy
    # Pre-filtro: ocultar filas donde G y H ambos son Fredy (sin problema)
    ws = wb["Hoja5"]
    num_filas = len(df_servicios)
    ws.conditional_formatting.add(
        f"H2:H{num_filas + 1}",
        FormulaRule(
            formula=['H2<>"Fredy Huaman R."'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        f"G2:G{num_filas + 1}",
        FormulaRule(
            formula=['G2<>"Fredy Huaman R."'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 2
    for r in range(2, ws.max_row + 1):
        g_val = ws.cell(row=r, column=7).value
        h_val = ws.cell(row=r, column=8).value
        if g_val == "Fredy Huaman R." and h_val == "Fredy Huaman R.":
            ws.row_dimensions[r].hidden = True

    # Hoja N° 6 — Nota Crédito Compensada
    # Pre-filtro: ocultar filas donde M<=0 y N=True (sin problema)
    ws = wb["Hoja6"]
    num_filas = len(hoja6)
    ws.conditional_formatting.add(
        f"M2:M{num_filas + 1}",
        FormulaRule(
            formula=['M2 > 0'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        f"N2:N{num_filas + 1}",
        FormulaRule(
            formula=['N2 = FALSE'],
            font=Font(color="FF0000"),
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid")
        )
    )
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 2
    for r in range(2, ws.max_row + 1):
        m_val = ws.cell(row=r, column=13).value or 0
        n_val = ws.cell(row=r, column=14).value
        if not (m_val > 0 or n_val is False or n_val == False):
            ws.row_dimensions[r].hidden = True

    # Hoja Datos Incompletos — formato
    ws = wb["Datos Incompletos"]
    num_filas_inc = len(df_incompletos)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 2
    # Cabecera naranja para identificar la hoja fácilmente
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="FF4500", fill_type="solid")
        cell.font = Font(name="Tahoma", size=9, bold=True, color="FFFFFFFF")
    # Marcar en amarillo las celdas de cada fila que están incompletas
    _col_indices = {name: idx + 1 for idx, name in enumerate(df_incompletos.columns) if name in _cols_validar}
    for r in range(2, num_filas_inc + 2):
        for col_name, col_idx in _col_indices.items():
            cell = ws.cell(row=r, column=col_idx)
            if _campo_incompleto(cell.value):
                cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                cell.font = Font(name="Tahoma", size=9, color="9C002A", bold=True)
            else:
                cell.font = Font(name="Tahoma", size=9)
        # Resto de celdas
        for cell in ws[r]:
            if cell.column not in _col_indices.values():
                cell.font = Font(name="Tahoma", size=9)
    ws.auto_filter.ref = ws.dimensions

    # Hoja Resumen Validaciones — formato
    ws = wb["Resumen Validaciones"]
    _col_colors = ['5B9BD5', 'A5A5A5', '70AD47', '4472C4']  # Validación, Descripción, Errores, Archivo
    for ci, color in enumerate(_col_colors, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = PatternFill(start_color=color, fill_type="solid")
        cell.font = Font(name="Tahoma", size=10, bold=True, color="FFFFFFFF")
    for r in range(2, len(df_resumen_val) + 2):
        errores = ws.cell(row=r, column=3).value or 0
        color_fila = "FFC7CE" if errores > 0 else "E2EFDA"
        font_err   = Font(name="Tahoma", size=10, bold=True, color="9C002A" if errores > 0 else "375623")
        for ci in range(1, 5):
            cell = ws.cell(row=r, column=ci)
            cell.font = font_err if ci == 3 else Font(name="Tahoma", size=10)
        ws.cell(row=r, column=3).fill = PatternFill(start_color=color_fila, fill_type="solid")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 62
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16

    # Mover "Resumen Validaciones" a la segunda posición (tras Hoja1)
    idx_actual = wb.sheetnames.index("Resumen Validaciones")
    wb.move_sheet("Resumen Validaciones", offset=-(idx_actual - 1))

    wb.save(nombre_archivo)

    # ============================================================================
    # PARTE ACTUALIZADA: CÁLCULO DE COMISIONES CON NUEVO MODELO
    # ============================================================================

    df = pd.read_excel('reporte.xlsx')
    df = df[df['Responsable 1'].notna()]
    endress = df[df['Producto CRM'].str.contains('Endress', na=False)]

    vendedores_db = list(database.vendedores_collection.find(
        {'esLider': {'$exists': False}}))
    umbrales = {v['nombre']: v['umbral_mensual'] for v in vendedores_db}
    uns = {v['nombre']: v['unidad_negocio'] for v in vendedores_db}

    trimestres = {
        'Q1': [1, 2, 3],
        'Q2': [4, 5, 6],
        'Q3': [7, 8, 9],
        'Q4': [10, 11, 12]
    }

    data = {}

    for _, meses in trimestres.items():
        result = (
            endress[endress['MES'].isin(meses)]
            .pivot_table(
                index='Responsable 1',
                columns='MES',
                values='Monto Total',
                aggfunc='sum',
                fill_value=0,
            )
            .assign(Total=lambda x: x.sum(axis=1))
            .reset_index()
        )
        result['Umbral Mensual'] = result['Responsable 1'].map(umbrales)
        result['Umbral Trimestral'] = result['Umbral Mensual'] * 3
        result['Unidad negocio'] = result['Responsable 1'].map(uns)
        result['Paso'] = result['Total'] > result['Umbral Trimestral']
        result = result[result['Unidad negocio'] == 'UNAU']
        for mes in meses:
            if mes in result.columns:
                result[f'Paso_{mes}'] = result['Paso'] | (
                    result[mes] > result['Umbral Mensual'])
            else:
                result[f'Paso_{mes}'] = result['Paso']
        for index, row in result.iterrows():
            for mes in meses:
                if mes in result.columns:
                    data[f'{mes}_{row["Responsable 1"]}'] = row[f'Paso_{mes}']

    df['MES_RESPONSABLE'] = df['MES'].astype(str) + '_' + df['Responsable 1']
    df['Comisiona'] = df['MES_RESPONSABLE'].map(data)
    df['Comisiona'] = np.where(df['Producto CRM'].str.contains(
        'Endress'), df['Comisiona'], True)
    df['Comisiona'] = df['Comisiona'].fillna(True)

    # Calcular comisión total base (1%)
    df['Comisión Total'] = np.where(
        df['Comisiona'], df['Monto Actualizado'] * 0.01, 0)

    # Calcular porcentajes individuales
    df['Porcentaje 1'] = 0.7
    df['Porcentaje 2'] = np.where(
        (df['Responsable 2'].str.contains('Paolo', na=False) &
         (df['Producto CRM'].str.contains('Proy', na=False))),
        0.5,
        0.3
    )

    # Calcular comisiones individuales (temporal para el reporte Excel)
    df['Comisión 1'] = df['Comisión Total'] * df['Porcentaje 1']
    df['Comisión 2'] = df['Comisión Total'] * df['Porcentaje 2']

    # ── Agregar Comisiona, Comisión 1, Comisión 2, Estado a Hoja1 ────────────
    wb2 = load_workbook(nombre_archivo)
    ws2 = wb2['Hoja1']

    # Headers (AC=29, AD=30, AE=31, AF=32)
    for col_num, header, color in [
        (29, 'Comisiona',  'A5A5A5'),
        (30, 'Comisión 1', '70AD47'),
        (31, 'Comisión 2', '70AD47'),
        (32, 'Estado',     'A5A5A5'),
    ]:
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color=color, fill_type='solid')
        cell.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

    # Fila a fila (idx 0-based → fila Excel = idx + 2)
    for idx, row in df.iterrows():
        r = idx + 2
        # Comisiona: valor directo
        c = ws2.cell(row=r, column=29)
        c.value = bool(row['Comisiona'])
        c.font = Font(name='Tahoma', size=9)
        # Comisión 1 y 2: fórmulas referenciando AC
        ws2[f'AD{r}'] = f'=IF(AC{r},S{r}*0.01*0.7,0)'
        ws2[f'AE{r}'] = (
            f'=IF(AC{r},'
            f'S{r}*0.01*IF(AND(ISNUMBER(SEARCH("Paolo",E{r})),ISNUMBER(SEARCH("Proy",K{r}))),0.5,0.3),'
            f'0)'
        )
        ws2[f'AD{r}'].font = Font(name='Tahoma', size=9)
        ws2[f'AE{r}'].font = Font(name='Tahoma', size=9)
        # Estado: valor directo
        c = ws2.cell(row=r, column=32)
        c.value = 'Estándar' if row['Comisiona'] else 'Atípico'
        c.font = Font(name='Tahoma', size=9)

    ws2.column_dimensions['AC'].width = 12
    ws2.column_dimensions['AD'].width = 14
    ws2.column_dimensions['AE'].width = 14
    ws2.column_dimensions['AF'].width = 12

    # ── HOJAS 7-10: Una hoja por trimestre – Umbral UNAU ─────────────────────
    NOMBRES_MESES = {
        1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr',
        5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago',
        9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
    }

    # Vendedores UNAU base (para mostrar incluso si no tuvieron ventas)
    vendedores_unau = [v for v, u in uns.items() if u == 'UNAU']

    def _escribir_hoja_trimestre(wb_obj, sheet_name, q_label, q_meses, df_endress):
        """Escribe las dos tablas de umbral para un trimestre en una hoja nueva."""
        endress_q = df_endress[df_endress['MES'].isin(q_meses)]

        # Pivot con todos los vendedores UNAU, incluyendo los que no tienen ventas
        if not endress_q.empty:
            pivot_q = (
                endress_q.pivot_table(
                    index='Responsable 1',
                    columns='MES',
                    values='Monto Total',
                    aggfunc='sum',
                    fill_value=0,
                )
                .reset_index()
            )
        else:
            pivot_q = pd.DataFrame({'Responsable 1': []})

        # Asegurar que todos los vendedores UNAU aparezcan
        base = pd.DataFrame({'Responsable 1': vendedores_unau})
        pivot_q = base.merge(pivot_q, on='Responsable 1', how='left').fillna(0)

        for m in q_meses:
            if m not in pivot_q.columns:
                pivot_q[m] = 0.0

        pivot_q['Total General'] = pivot_q[[m for m in q_meses]].sum(axis=1)
        pivot_q['Umbral Mensual'] = pivot_q['Responsable 1'].map(umbrales).fillna(0)
        pivot_q['Umbral Trimestral'] = pivot_q['Umbral Mensual'] * 3
        pivot_q['Paso'] = pivot_q.apply(
            lambda row: row['Total General'] > row['Umbral Trimestral'] or
                        any(row[m] > row['Umbral Mensual'] for m in q_meses),
            axis=1
        )

        ws = wb_obj.create_sheet(sheet_name)

        # Título
        title_cell = ws['A1']
        title_cell.value = f'Resumen Umbral Trimestral – {q_label}'
        title_cell.font = Font(name='Tahoma', size=11, bold=True)

        # ── Tabla 1: montos por mes ───────────────────────────────────────────
        T1_ROW = 3
        t1_headers = ['Vendedor'] + [NOMBRES_MESES[m] for m in q_meses] + ['Total General']
        for ci, hdr in enumerate(t1_headers, 1):
            cell = ws.cell(row=T1_ROW, column=ci)
            cell.value = hdr
            cell.fill = PatternFill(start_color='5B9BD5', fill_type='solid')
            cell.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

        for ri, (_, row) in enumerate(pivot_q.iterrows(), 1):
            r = T1_ROW + ri
            ws.cell(row=r, column=1).value = row['Responsable 1']
            ws.cell(row=r, column=1).font = Font(name='Tahoma', size=9)
            for ci, m in enumerate(q_meses, 2):
                cell = ws.cell(row=r, column=ci)
                cell.value = float(row.get(m, 0))
                cell.number_format = '#,##0.00'
                cell.font = Font(name='Tahoma', size=9)
            tc = ws.cell(row=r, column=len(q_meses) + 2)
            tc.value = float(row['Total General'])
            tc.number_format = '#,##0.00'
            tc.font = Font(name='Tahoma', size=9, bold=True)

        # ── Tabla 2: umbral y resultado ───────────────────────────────────────
        T2_ROW = T1_ROW + len(pivot_q) + 3

        # Subtítulo tabla 2
        sub = ws.cell(row=T2_ROW - 1, column=1)
        sub.value = '¿Pasó el umbral?'
        sub.font = Font(name='Tahoma', size=9, bold=True, italic=True)

        t2_headers = ['Vendedor', 'Umbral Mensual', 'Umbral Trimestral', 'Total', '¿Pasó?']
        t2_colors = ['A5A5A5', 'FFC000', 'FFC000', '5B9BD5', '70AD47']
        for ci, (hdr, color) in enumerate(zip(t2_headers, t2_colors), 1):
            cell = ws.cell(row=T2_ROW, column=ci)
            cell.value = hdr
            cell.fill = PatternFill(start_color=color, fill_type='solid')
            cell.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

        for ri, (_, row) in enumerate(pivot_q.iterrows(), 1):
            r = T2_ROW + ri
            ws.cell(row=r, column=1).value = row['Responsable 1']
            ws.cell(row=r, column=1).font = Font(name='Tahoma', size=9)

            for ci, col_key, fmt in [
                (2, 'Umbral Mensual', '#,##0.00'),
                (3, 'Umbral Trimestral', '#,##0.00'),
                (4, 'Total General', '#,##0.00'),
            ]:
                cell = ws.cell(row=r, column=ci)
                cell.value = float(row[col_key])
                cell.number_format = fmt
                cell.font = Font(name='Tahoma', size=9)

            paso = bool(row['Paso'])
            paso_cell = ws.cell(row=r, column=5)
            paso_cell.value = 'SI' if paso else 'NO'
            paso_cell.fill = PatternFill(
                start_color='70AD47' if paso else 'FF4040', fill_type='solid')
            paso_cell.font = Font(name='Tahoma', size=9, bold=True, color='FFFFFFFF')

        # Ajustar anchos
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    mes_actual = pd.Timestamp.now().month
    trimestres_a_generar = [
        ('Hoja7',  'Q1 (Ene–Mar)', [1, 2, 3]),
        ('Hoja8',  'Q2 (Abr–Jun)', [4, 5, 6]),
        ('Hoja9',  'Q3 (Jul–Sep)', [7, 8, 9]),
        ('Hoja10', 'Q4 (Oct–Dic)', [10, 11, 12]),
    ]
    for sheet_name, q_label, q_meses in trimestres_a_generar:
        if q_meses[0] <= mes_actual:
            _escribir_hoja_trimestre(wb2, sheet_name, q_label, q_meses, endress)
    # ─────────────────────────────────────────────────────────────────────────

    wb2.save(nombre_archivo)
    # ─────────────────────────────────────────────────────────────────────────

    # Determinar producto simplificado
    df['Producto'] = np.where(
        df['Producto CRM'].str.contains('Endress', na=False),
        'Endress',
        df['Producto CRM']
    )

    df.drop(columns=['MES_RESPONSABLE'], inplace=True)

    # ============================================================================
    # CREAR ARRAY DE RESPONSABLES (NUEVO MODELO) Y CONSOLIDAR DUPLICADOS
    # ============================================================================

    def crear_array_responsables(row):
        """
        Crea el array de responsables consolidando duplicados
        """
        responsables_dict = {}

        # Responsable 1
        if pd.notna(row['Responsable 1']) and row['Responsable 1'].strip() != '':
            nombre_r1 = row['Responsable 1'].strip()
            nombre_key_r1 = nombre_r1.lower()

            responsables_dict[nombre_key_r1] = {
                'nombre': nombre_r1,
                'porcentaje': row['Porcentaje 1'],
                'comision': row['Comisión 1']
            }

        # Responsable 2
        if pd.notna(row['Responsable 2']) and row['Responsable 2'].strip() != '':
            nombre_r2 = row['Responsable 2'].strip()
            nombre_key_r2 = nombre_r2.lower()

            if nombre_key_r2 in responsables_dict:
                # Es el mismo responsable, consolidar
                responsables_dict[nombre_key_r2]['porcentaje'] += row['Porcentaje 2']
                responsables_dict[nombre_key_r2]['comision'] += row['Comisión 2']
            else:
                # Es diferente responsable
                responsables_dict[nombre_key_r2] = {
                    'nombre': nombre_r2,
                    'porcentaje': row['Porcentaje 2'],
                    'comision': row['Comisión 2']
                }

        # Convertir dict a lista
        return list(responsables_dict.values())

    # Aplicar la función para crear el array de responsables
    df['responsables'] = df.apply(crear_array_responsables, axis=1)

    # ============================================================================
    # MAPEO DE COLUMNAS PARA MONGODB
    # ============================================================================

    COLUMN_MAP = {
        "OK": "ok",
        "AÑO": "anio",
        "MES": "mes",
        "Unidad de Negocio": "unidad_negocio",
        "Fecha": "fecha",
        "Estado": "estado",
        "Número": "numero",
        "Monto Total": "monto_total",
        "Producto CRM": "producto_crm",
        "UBruta": "utilidad_bruta",
        "Nombre Empresa": "nombre_empresa",
        "Subject": "subject",
        "Codigos": "codigos",
        "Cotizacion #": "cotizacion_num",
        "Proviene EPC/OEM/Canal Deal?": "origen_deal",
        "T/C de la Factura": "tipo_cambio_factura",
        "Monto Actualizado": "monto_actualizado",
        "Diferencia": "diferencia",
        "Notas": "notas",
        "Observaciones": "observaciones",
        "Periodo": "periodo",
        "EstadoPago-Vendedor": "estado_pago_vendedor",
        "Lider 1": "lider_1",
        "Lider 2": "lider_2",
        "EstadoPago-Lideres": "estado_pago_lideres",
        "Umbral": "umbral",
        "Producto": "producto",
        "Comisiona": "comisiona",
        "Comisión Total": "comision_total",
        "responsables": "responsables"  # Nuevo campo
    }

    # Renombrar columnas
    df = df.rename(columns=COLUMN_MAP)

    # Seleccionar solo las columnas que vamos a insertar
    columnas_a_insertar = [
        "ok", "anio", "mes", "unidad_negocio", "fecha", "estado", "numero",
        "monto_total", "producto_crm", "utilidad_bruta", "nombre_empresa",
        "subject", "codigos", "cotizacion_num", "origen_deal",
        "tipo_cambio_factura", "monto_actualizado", "diferencia", "notas",
        "observaciones", "periodo", "estado_pago_vendedor", "lider_1",
        "lider_2", "estado_pago_lideres", "umbral", "producto", "comisiona",
        "comision_total", "responsables"
    ]

    df = df[columnas_a_insertar]

    # Convertir a registros
    records = df.to_dict(orient="records")

    # Preservar responsables editados manualmente antes de borrar
    facturas_manuales = {
        doc['numero']: [
            {'nombre': r['nombre'], 'porcentaje': r['porcentaje']}
            for r in doc.get('responsables', [])
        ]
        for doc in database.invoices_collection.find(
            {'manually_edited': True},
            {'numero': 1, 'responsables': 1}
        )
    }

    # Limpiar e insertar
    database.invoices_collection.delete_many({})
    database.invoices_collection.insert_many(records)

    # Restaurar responsables editados manualmente, recalculando comisiones con los nuevos montos
    if facturas_manuales:
        numeros_editados = list(facturas_manuales.keys())
        for doc in database.invoices_collection.find(
            {'numero': {'$in': numeros_editados}},
            {'numero': 1, 'comision_total': 1, 'comisiona': 1}
        ):
            numero = doc['numero']
            comision_total = doc.get('comision_total', 0)
            comisiona = doc.get('comisiona', True)

            nuevos_responsables = []
            for r in facturas_manuales[numero]:
                nuevos_responsables.append({
                    'nombre': r['nombre'],
                    'porcentaje': r['porcentaje'],
                    'comision': comision_total * r['porcentaje'] if comisiona else 0
                })

            database.invoices_collection.update_one(
                {'numero': numero},
                {'$set': {'responsables': nuevos_responsables, 'manually_edited': True}}
            )

    # Actualizar reporte.xlsx con los cambios manuales restaurados
    if facturas_manuales:
        wb_excel = load_workbook(nombre_archivo)
        ws_hoja1 = wb_excel['Hoja1']

        # Construir mapa Número → fila Excel (col I = 9, datos desde fila 2)
        numero_a_fila = {}
        for row_idx in range(2, ws_hoja1.max_row + 1):
            val = ws_hoja1.cell(row=row_idx, column=9).value
            if val is not None:
                numero_a_fila[str(val)] = row_idx

        font_celda = Font(name='Tahoma', size=9)

        for doc in database.invoices_collection.find(
            {'numero': {'$in': list(facturas_manuales.keys())}},
            {'numero': 1, 'responsables': 1}
        ):
            numero = doc['numero']
            fila = numero_a_fila.get(str(numero))
            if fila is None:
                continue

            responsables = doc.get('responsables', [])

            # Responsable 1 (col D=4)
            r1_nombre = responsables[0]['nombre'] if len(responsables) >= 1 else None
            r1_comision = responsables[0]['comision'] if len(responsables) >= 1 else 0
            c = ws_hoja1.cell(row=fila, column=4)
            c.value = r1_nombre
            c.font = font_celda

            # Responsable 2 (col E=5)
            r2_nombre = responsables[1]['nombre'] if len(responsables) >= 2 else None
            r2_comision = responsables[1]['comision'] if len(responsables) >= 2 else 0
            c = ws_hoja1.cell(row=fila, column=5)
            c.value = r2_nombre
            c.font = font_celda

            # Reemplazar fórmulas de Comisión 1 (AD) y Comisión 2 (AE) con valores reales
            c_ad = ws_hoja1.cell(row=fila, column=30)
            c_ad.value = r1_comision
            c_ad.font = font_celda

            c_ae = ws_hoja1.cell(row=fila, column=31)
            c_ae.value = r2_comision
            c_ae.font = font_celda

        wb_excel.save(nombre_archivo)
        print(f"✅ reporte.xlsx actualizado con {len(facturas_manuales)} factura(s) editada(s) manualmente")

    print(f"\n{'=' * 80}")
    print(f"✅ DATOS CARGADOS A MONGODB")
    print(f"{'=' * 80}")
    print(f"Total de facturas insertadas: {len(records)}")

    # Estadísticas de consolidación
    facturas_consolidadas = sum(1 for r in records if len(r['responsables']) == 1 and
                                r['responsables'][0]['porcentaje'] == 1.0)
    print(f"Facturas con responsables consolidados: {facturas_consolidadas}")
    print(f"{'=' * 80}\n")
