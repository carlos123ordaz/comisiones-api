from io import BytesIO
from bson import ObjectId
from typing import List, Optional
from openpyxl import load_workbook
from config.database import vendedores_collection
from models.vendedor import VendedorCreate, VendedorUpdate, Vendedor


def calcular_umbrales(meta_mensual: float, porcentaje_umbral: float):
    umbral_mensual = meta_mensual * porcentaje_umbral / 100
    return {
        "umbral_mensual": umbral_mensual,
        "umbral_trimestral": umbral_mensual * 3,
        "umbral_meta": umbral_mensual * 5/4
    }


def get_all_vendedores() -> List[dict]:
    vendedores = list(vendedores_collection.find(
        {'esLider': {'$exists': False}}))

    result = []
    for v in vendedores:
        result.append({
            'id': str(v['_id']),
            'nombre': v['nombre'],
            'meta_mensual': v['meta_mensual'],
            'porcentaje_umbral': v['porcentaje_umbral'],
            'unidad_negocio': v['unidad_negocio'],
            'umbral_mensual': v['umbral_mensual'],
            'umbral_trimestral': v['umbral_trimestral'],
            'umbral_meta': v['umbral_meta']
        })

    return result


def create_vendedor(vendedor: VendedorCreate) -> dict:
    existente = vendedores_collection.find_one({'nombre': vendedor.nombre})
    if existente:
        raise ValueError("El vendedor ya existe")

    umbrales = calcular_umbrales(
        vendedor.meta_mensual, vendedor.porcentaje_umbral)

    vendedor_dict = vendedor.model_dump()
    vendedor_dict.update(umbrales)
    result = vendedores_collection.insert_one(vendedor_dict)
    vendedor_dict['id'] = str(result.inserted_id)

    return vendedor_dict


def update_vendedor(vendedor_id: str, vendedor: VendedorUpdate) -> dict:
    vendedor_actual = vendedores_collection.find_one(
        {'_id': ObjectId(vendedor_id)})
    if not vendedor_actual:
        raise ValueError("Vendedor no encontrado")

    update_data = {}

    if vendedor.meta_mensual is not None:
        update_data['meta_mensual'] = vendedor.meta_mensual

    if vendedor.porcentaje_umbral is not None:
        update_data['porcentaje_umbral'] = vendedor.porcentaje_umbral

    if vendedor.unidad_negocio is not None:
        update_data['unidad_negocio'] = vendedor.unidad_negocio

    if 'meta_mensual' in update_data or 'porcentaje_umbral' in update_data:
        meta = update_data.get(
            'meta_mensual', vendedor_actual['meta_mensual'])
        porcentaje = update_data.get(
            'porcentaje_umbral', vendedor_actual['porcentaje_umbral'])
        umbrales = calcular_umbrales(meta, porcentaje)
        update_data.update(umbrales)

    vendedores_collection.update_one(
        {'_id': ObjectId(vendedor_id)},
        {'$set': update_data}
    )

    vendedor_actualizado = vendedores_collection.find_one(
        {'_id': ObjectId(vendedor_id)})
    vendedor_actualizado['id'] = str(vendedor_actualizado['_id'])

    return vendedor_actualizado


def delete_vendedor(vendedor_id: str) -> bool:
    result = vendedores_collection.delete_one(
        {'_id': ObjectId(vendedor_id)})

    if result.deleted_count == 0:
        raise ValueError("Vendedor no encontrado")

    return True


def importar_vendedores_excel(file_bytes: bytes) -> dict:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required = ['RESPONSIBLE_ID', 'DEPARTAMENTO', 'META MENSUAL ($)', 'Umbral']
    for col in required:
        if col not in headers:
            raise ValueError(f"Columna requerida no encontrada: '{col}'")

    col_idx = {h: i for i, h in enumerate(headers)}
    creados = 0
    actualizados = 0
    errores = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nombre = row[col_idx['RESPONSIBLE_ID']]
        if not nombre:
            continue

        nombre = str(nombre).strip()
        departamento = str(row[col_idx['DEPARTAMENTO']] or '-').strip()
        meta_mensual = row[col_idx['META MENSUAL ($)']]
        umbral_pct = row[col_idx['Umbral']]

        if meta_mensual is None or umbral_pct is None:
            errores.append(f"Fila {row_num}: datos incompletos para '{nombre}'")
            continue

        try:
            meta_mensual = float(meta_mensual)
            umbral_pct = float(umbral_pct)
        except (ValueError, TypeError):
            errores.append(f"Fila {row_num}: valores numéricos inválidos para '{nombre}'")
            continue

        # El Excel tiene el umbral como decimal (0.6 = 60%), convertir a porcentaje
        if umbral_pct <= 1:
            umbral_pct = umbral_pct * 100

        umbrales = calcular_umbrales(meta_mensual, umbral_pct)

        existente = vendedores_collection.find_one({'nombre': nombre})
        if existente:
            vendedores_collection.update_one(
                {'_id': existente['_id']},
                {'$set': {
                    'meta_mensual': meta_mensual,
                    'porcentaje_umbral': umbral_pct,
                    'unidad_negocio': departamento,
                    **umbrales
                }}
            )
            actualizados += 1
        else:
            vendedores_collection.insert_one({
                'nombre': nombre,
                'meta_mensual': meta_mensual,
                'porcentaje_umbral': umbral_pct,
                'unidad_negocio': departamento,
                **umbrales
            })
            creados += 1

    return {
        'creados': creados,
        'actualizados': actualizados,
        'errores': errores,
        'message': f'{creados} creados, {actualizados} actualizados' + (f', {len(errores)} errores' if errores else '')
    }


def get_vendedor_by_nombre(nombre: str) -> Optional[dict]:
    return vendedores_collection.find_one({'nombre': nombre})


def get_usuarios_info() -> List[dict]:
    vendedores = list(vendedores_collection.find(
        {'esLider': {'$exists': False}}).sort('nombre', 1))

    usuarios_info = []
    for vendedor in vendedores:
        info = {
            'nombre': vendedor['nombre'],
            'unidad_negocio': vendedor['unidad_negocio'],
            'umbral_mensual': vendedor['umbral_mensual']
        }
        usuarios_info.append(info)

    return usuarios_info
